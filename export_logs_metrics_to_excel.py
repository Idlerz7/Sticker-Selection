#!/usr/bin/env python3
"""
Scan logs/ for training (train.log) and test (test_*.log) entries, parse
[StructuredEval] / [EvalSummary] metrics lines, write an .xlsx workbook.

Requires: pip install openpyxl
"""

from __future__ import annotations

import argparse
import os
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

RE_METRIC_LINE = re.compile(
    r"\[(StructuredEval|EvalSummary)\] epoch=(\d+) mode=(\S+) total=([\d.]+) "
    r"r@1=([\d.]+) r@2=([\d.]+) r@5=([\d.]+) mrr=([\d.]+)"
)
RE_TEST_SUMMARY = re.compile(
    r"\[StructuredTestSummary\] samples=(\d+)\s+metrics_file=(\S+)\s+pred_file=(\S+)"
)


def parse_log_file(path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Return (metric_rows, test_summary_rows) from one file."""
    metrics: List[Dict[str, Any]] = []
    summaries: List[Dict[str, Any]] = []
    text = path.read_text(encoding="utf-8", errors="replace")
    for line in text.splitlines():
        ts = ""
        if len(line) > 23 and line[4] == "-" and line[7] == "-":
            ts = line[:23].strip()

        m = RE_METRIC_LINE.search(line)
        if m:
            kind, ep, mode, tot, r1, r2, r5, mrr = m.groups()
            metrics.append(
                {
                    "timestamp": ts,
                    "line_kind": kind,
                    "epoch": int(ep),
                    "mode": mode,
                    "total": float(tot),
                    "r@1": float(r1),
                    "r@2": float(r2),
                    "r@5": float(r5),
                    "mrr": float(mrr),
                }
            )
            continue

        m2 = RE_TEST_SUMMARY.search(line)
        if m2:
            summaries.append(
                {
                    "timestamp": ts,
                    "samples": int(m2.group(1)),
                    "metrics_file": m2.group(2),
                    "pred_file": m2.group(3),
                }
            )
    return metrics, summaries


def dedupe_training_rows(
    rows: List[Dict[str, Any]], log_path: str
) -> List[Dict[str, Any]]:
    """
    Same epoch may appear twice (e.g. sanity total=2 then real total=3542).
    Keep the row with largest `total` per (log_path, epoch, mode).
    """
    best: Dict[Tuple[str, int, str], Dict[str, Any]] = {}
    for r in rows:
        key = (log_path, r["epoch"], r["mode"])
        prev = best.get(key)
        if prev is None or r["total"] > prev["total"]:
            best[key] = r
    out = list(best.values())
    out.sort(key=lambda x: (x["epoch"], x["mode"]))
    return out


def collect_logs(logs_root: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    training_out: List[Dict[str, Any]] = []
    test_out: List[Dict[str, Any]] = []

    for dirpath, _, filenames in os.walk(logs_root):
        for fn in filenames:
            if not fn.endswith(".log"):
                continue
            path = Path(dirpath) / fn
            rel = str(path.relative_to(logs_root))
            metrics, summaries = parse_log_file(path)

            is_train_log = fn == "train.log"
            is_test_log = fn.startswith("test_")

            if is_train_log and metrics:
                clean = dedupe_training_rows(metrics, rel)
                for r in clean:
                    training_out.append(
                        {
                            "log_file": rel,
                            "timestamp": r["timestamp"],
                            "metric_tag": r["line_kind"],
                            "epoch": r["epoch"],
                            "mode": r["mode"],
                            "total": r["total"],
                            "r@1": r["r@1"],
                            "r@2": r["r@2"],
                            "r@5": r["r@5"],
                            "mrr": r["mrr"],
                        }
                    )

            if is_test_log and metrics:
                summ = summaries[-1] if summaries else {}
                for r in metrics:
                    test_out.append(
                        {
                            "log_file": rel,
                            "timestamp": r["timestamp"],
                            "metric_tag": r["line_kind"],
                            "epoch": r["epoch"],
                            "mode": r["mode"],
                            "total": r["total"],
                            "r@1": r["r@1"],
                            "r@2": r["r@2"],
                            "r@5": r["r@5"],
                            "mrr": r["mrr"],
                            "structured_samples": summ.get("samples", "") if summ else "",
                            "metrics_json": summ.get("metrics_file", ""),
                            "pred_json": summ.get("pred_file", ""),
                        }
                    )

    training_out.sort(key=lambda x: (x["log_file"], x["epoch"], x["mode"]))
    test_out.sort(key=lambda x: (x["log_file"], x["epoch"]))
    return training_out, test_out


def write_xlsx(
    training: List[Dict[str, Any]],
    test_rows: List[Dict[str, Any]],
    out_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "training_epochs"
    if training:
        headers = list(training[0].keys())
        ws_t.append(headers)
        for row in training:
            ws_t.append([row[h] for h in headers])
    else:
        ws_t.append(["(no training metric lines found)"])

    ws_te = wb.create_sheet("test_runs")
    if test_rows:
        headers = list(test_rows[0].keys())
        ws_te.append(headers)
        for row in test_rows:
            ws_te.append([row[h] for h in headers])
    else:
        ws_te.append(["(no test_*.log metric lines found)"])

    for ws in (ws_t, ws_te):
        if ws.max_column < 1:
            continue
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            hdr = str(ws.cell(1, col).value or "")
            ws.column_dimensions[letter].width = min(48, max(10, len(hdr) + 2))

    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--logs-root",
        type=Path,
        default=Path("logs"),
        help="Root directory to scan (default: ./logs)",
    )
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: logs/metrics_export_YYYYMMDD_HHMMSS.xlsx)",
    )
    args = ap.parse_args()

    root = args.logs_root.resolve()
    if not root.is_dir():
        raise SystemExit(f"Not a directory: {root}")

    out = args.output
    if out is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = root / f"metrics_export_{stamp}.xlsx"

    training, test_rows = collect_logs(root)
    write_xlsx(training, test_rows, out.resolve())
    print(f"Wrote {out}  (training rows={len(training)}, test rows={len(test_rows)})")


if __name__ == "__main__":
    main()
